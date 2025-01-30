from django.shortcuts import render,get_object_or_404,redirect
from django.views import View
from .models import *
from django.http import JsonResponse,HttpResponse
import os,mimetypes,json,openpyxl,re,pytz
from django.contrib.auth.models import User
from django.conf import settings
from django.core.mail import EmailMessage,send_mail
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from html import unescape
from html.parser import HTMLParser
from openpyxl.styles import Font
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings
from django.core.mail import EmailMultiAlternatives

from email.header import decode_header
from email.utils import make_msgid
from django.core.files.storage import default_storage
from django.db.models import Max, Count, Q,F
from django.db.models import Subquery, OuterRef
from datetime import datetime, timedelta
from django.utils.timezone import localtime
from django.template.loader import render_to_string
from django.conf import settings
from acsapp.models import *

def mailtemplate(req):
    return render(req,"mail_template.html")





def group_detail(request, group_id):
    groups = get_object_or_404(Group, id=group_id)
    current_user = request.user  

    if current_user != groups.created_by and current_user not in groups.members.all():
        return HttpResponse("You are not in the group") 

    grouped_messages = {}

    # Group messages by date
    messages_in_group = GroupMessage.objects.filter(group=groups).order_by('created_at')
    for message in messages_in_group:
        date = message.created_at.date()
        if date not in grouped_messages:
            grouped_messages[date] = []
        grouped_messages[date].append(message)

    # Fetch last message details for all groups the current user is a member of or created
    last_message_subquery = GroupMessage.objects.filter(
        group=OuterRef('id')
    ).order_by('-created_at')
    last_file_subquery = GroupMessage.objects.filter(
        group=OuterRef('id'),
        sender=current_user
    ).order_by('-created_at').values('file')[:1]

    # Filter groups that the current user is either a member of or created
    all_groups = Group.objects.filter(
        Q(created_by=current_user) | Q(members=current_user)
    ).annotate(
        last_message=Subquery(last_message_subquery.values('content')[:1]),
        last_message_time=Subquery(last_message_subquery.values('created_at')[:1]),
        last_message_sender=Subquery(last_message_subquery.values('sender__username')[:1]),
        last_sent_file=Subquery(last_file_subquery)  # Add the last sent file to the annotation
    ).distinct()

    users = User.objects.exclude(id=current_user.id).annotate(
        last_message_content=Subquery(
            Message.objects.filter(
                receiver=OuterRef('id'),
                sender=current_user
            ).order_by('-created_at').values('content')[:1]
        ),
        last_message_time=Subquery(
            Message.objects.filter(
                receiver=OuterRef('id'),
                sender=current_user
            ).order_by('-created_at').values('created_at')[:1]
        ),
        last_message_sender=Subquery(
            Message.objects.filter(
                receiver=OuterRef('id'),
                sender=current_user
            ).order_by('-created_at').values('sender__username')[:1]
        ),
        unread_count=Count('received_messages', filter=Q(received_messages__is_read=False, received_messages__sender=current_user))  # Unread message count
    ).order_by('-last_message_time')  # Sort users by the most recent message time

    # Fetch group members and creator
    group_members = groups.members.all()
    group_creator = groups.created_by

    # Fetch the last message sent by the current user in the selected group
    last_message_by_user = GroupMessage.objects.filter(
        group=groups,
        sender=current_user
    ).order_by('-created_at').first()

    context = {
        'allgroups': all_groups,
        'groups': groups,
        'group_members': group_members,
        'group_creator': group_creator,
        'grouped_messages': grouped_messages,
        'users': users,
    }

    return render(request, 'sms.html', context)

@login_required
def send_message(request, recipient_id=None, group_id=None):

    print("send_message called")
    print("User clikcm")
    current_user = request.user
    recipient_online_status = None

    recipient = None
    groupname = None
    group_members = None
    groups = None
  

    # Handle individual messaging
    if recipient_id:
        recipient = get_object_or_404(User, id=recipient_id)
           # Check if the recipient is online
        recipient_online_status = chatLoginHistory.objects.filter(
            user=recipient,
            logout_time__isnull=True
        ).exists()
        print(recipient_online_status,"status")
    
   # Subquery to fetch the last message details for each group
    last_message_subquery = GroupMessage.objects.filter(
        group=OuterRef('id')
    ).order_by('-created_at')
    last_file_subquery = GroupMessage.objects.filter(
        group=OuterRef('id'),
        sender=current_user
    ).order_by('-created_at').values('file')[:1]

  


    # Filter groups that the current user is either a member of or created
    allgroups = Group.objects.filter(
        Q(created_by=current_user) | Q(members=current_user)
    ).annotate(
        last_message=Subquery(last_message_subquery.values('content')[:1]),
        last_message_time=Subquery(last_message_subquery.values('created_at')[:1]),
        last_message_sender=Subquery(last_message_subquery.values('sender__username')[:1]),
        last_sent_file=Subquery(last_file_subquery)  # Add the last sent file to the annotation
    ).distinct()

    # print(allgroups,"all grpssss")


    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        uploaded_file = request.FILES.get('file')
        audio_message = request.POST.get('audio_message')

        if not content and not uploaded_file and not audio_message:
            messages.error(request, "Please provide a message or upload a file.")
        else:
            # Individual message logic
            if recipient:
                parent_message_id = request.POST.get('parent_message_id')
                parent_message = Message.objects.filter(id=parent_message_id).first() if parent_message_id else None

                message = Message.objects.create(
                    sender=current_user,
                    receiver=recipient,
                    content=content or None,
                    parent_message=parent_message
                )
                message.is_delivered = True

                if uploaded_file:
                    file_path = default_storage.save(uploaded_file.name, uploaded_file)
                    message.file = file_path

                message.save()
 # Create a notification for the receiver
                chatNotification.objects.create(
                    user=recipient,
                    message=message,
                    is_read=False
                )
              

    # Fetch messages for rendering
    # grouped_messages = {}
    if recipient:
        messages = Message.objects.filter(
            Q(sender=current_user, receiver=recipient) |
            Q(sender=recipient, receiver=current_user)
        ).order_by('created_at')

        categorized_messages = categorize_messages_by_date(messages)
        grouped_messages = group_messages_by_date(categorized_messages)

        # Mark messages as read if the recipient views them
        unread_messages = messages.filter(receiver=current_user, is_read=False)
        unread_messages.update(is_read=True)

  

    # Fetch users excluding the current user for the sidebar
    last_messages = Message.objects.filter(
        Q(sender=OuterRef('id'), receiver=current_user) |
        Q(sender=current_user, receiver=OuterRef('id'))
    ).order_by('-created_at')

    users = User.objects.exclude(id=current_user.id).annotate(
        last_message_time=Subquery(last_messages.values('created_at')[:1]),
        last_message_content=Subquery(last_messages.values('content')[:1]),
        last_message_is_read=Subquery(last_messages.values('is_read')[:1]),
    ).order_by('-last_message_time')
    
    groupname1 = Group.objects.all()
    groupname2 = Group.objects.filter(members=request.user)
    print(allgroups,"All grps")
    # print(user_groupsvinoth,"User grps")

    return render(request, 'sms.html', {
        'recipient': recipient,
        'grouped_messages': grouped_messages,
        'users': users,
        'current_user': current_user,
        'groups': groups,
        'group_members': group_members,
        # "groupname1":groupname1,
        "groupname2":groupname2,
        "allgroups":allgroups,
        'recipient': recipient if recipient_id else None,
        'recipient_online_status': recipient_online_status,
    })


def categorize_messages_by_date(messages):
    today = localtime().date()
    yesterday = today - timedelta(days=1)

    categorized_messages = []
    for message in messages:
        msg_date = localtime(message.created_at).date()
        if msg_date == today:
            category = "Today"
        elif msg_date == yesterday:
            category = "Yesterday"
        else:
            category = msg_date.strftime("%d-%b-%Y")

        categorized_messages.append((category, message))

    return categorized_messages


def group_messages_by_date(categorized_messages):
    grouped_messages = {}
    for category, message in categorized_messages:
        grouped_messages.setdefault(category, []).append(message)

    return grouped_messages


def send_group_message(request, group_id):
    current_user = request.user  
    groupname1 = Group.objects.all()
    groups = get_object_or_404(Group, id=group_id)
    group = Group.objects.get(id=group_id)
    group_members = group.members.all()  
    group_creator = group.created_by  

   # Subquery to fetch the last message details for each group
    last_message_subquery = GroupMessage.objects.filter(
        group=OuterRef('id')
    ).order_by('-created_at')

    

    # Filter groups that the current user is either a member of or created
    allgroups = Group.objects.filter(
        Q(created_by=current_user) | Q(members=current_user)
    ).annotate(
        last_message=Subquery(last_message_subquery.values('content')[:1]),
        last_message_time=Subquery(last_message_subquery.values('created_at')[:1]),
        last_message_sender=Subquery(last_message_subquery.values('sender__username')[:1]),
    ).distinct()


    


 # Get users excluding the current user and annotate with last message content, time, and unread count
    users = User.objects.exclude(id=current_user.id).annotate(
        last_message_content=Subquery(
            Message.objects.filter(
                receiver=OuterRef('id'),
                sender=current_user
            ).order_by('-created_at').values('content')[:1]
        ),
        last_message_time=Subquery(
            Message.objects.filter(
                receiver=OuterRef('id'),
                sender=current_user
            ).order_by('-created_at').values('created_at')[:1]
        ),
        last_message_sender=Subquery(
            Message.objects.filter(
                receiver=OuterRef('id'),
                sender=current_user
            ).order_by('-created_at').values('sender__username')[:1]
        ),
        unread_count=Count('received_messages', filter=Q(received_messages__is_read=False, received_messages__sender=current_user))  # Unread message count
    ).order_by('-last_message_time')  # Sort users by the most recent message time
    if request.method == 'POST':
        content = request.POST.get('content')
        file = request.FILES.get('file')
        print(content, "grp content")
        parent_message_id = request.POST.get('parent_message_id')

        if not content and not file:
            return HttpResponse("Please provide a message or select a file.")
        
        # Get the parent message if it exists
        parent_message = None
        if parent_message_id:
            parent_message = get_object_or_404(GroupMessage, id=parent_message_id)

        group_message = GroupMessage.objects.create(
            group=group,
            sender=request.user,
            content=content,
            is_individual=False,
            parent_message=parent_message
        )

        if file:
            group_message.file = file
            group_message.save()

    # Fetch the grouped messages to display
    grouped_messages = {}
    messages_in_group = GroupMessage.objects.filter(group=group, is_individual=False).order_by('created_at')

    # Group messages by date
    for message in messages_in_group:
        date = message.created_at.date()
        if date not in grouped_messages:
            grouped_messages[date] = []
        grouped_messages[date].append(message)

    return render(request, 'sms.html', {
        'group': group,
        'grouped_messages': grouped_messages,
        'users': users, 
        # 'groupname1': groupname1,
        'groups': groups,
        'group_members': group_members,
        'group_creator': group_creator,
        'allgroups': allgroups,
    })

    # Render the page for GET requests
  




@login_required
def check_message_status(request):
    messages = Message.objects.filter(receiver=request.user, is_delivered=True).values('id', 'is_read', 'is_delivered')
    return JsonResponse({'messages': list(messages)})
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from .models import Message, GroupMessage

@login_required
def get_notifications(request):
    notifications = []
    current_user = request.user

    # Fetch unread messages from individual users
    unread_user_messages = Message.objects.filter(receiver=current_user, is_read=False)
    unread_user_counts = unread_user_messages.values('sender').annotate(unread_count=Count('id'))

    for unread in unread_user_counts:
        notifications.append({
            'id': f'user-{unread["sender"]}',
            'user_id': unread['sender'],
            'unread_count': unread['unread_count'],
            'type': 'user',
        })

    # Fetch unread messages from groups
    unread_group_messages = GroupMessage.objects.filter(group__members=current_user).exclude(sender=current_user).filter(is_read=False)
    unread_group_counts = unread_group_messages.values('group').annotate(unread_count=Count('id'))

    for unread in unread_group_counts:
        notifications.append({
            'id': f'group-{unread["group"]}',
            'group_id': unread['group'],
            'unread_count': unread['unread_count'],
            'type': 'group',
        })

    return JsonResponse({'notifications': notifications})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Message, GroupMessage
import json

@csrf_exempt
@login_required
def mark_notifications_as_read(request):
    if request.method == "POST":
        data = json.loads(request.body)
        notification_id = data.get('notification_id')

        if notification_id:
            if notification_id.startswith("user-"):
                user_id = notification_id.split("-")[1]
                Message.objects.filter(sender_id=user_id, receiver=request.user, is_read=False).update(is_read=True)
            elif notification_id.startswith("group-"):
                group_id = notification_id.split("-")[1]
                GroupMessage.objects.filter(group_id=group_id, group__members=request.user, is_read=False).update(is_read=True)

            return JsonResponse({"success": True})

    return JsonResponse({"success": False}, status=400)




@csrf_exempt
def mark_message_as_read(request):
    if request.method == "POST":
        message_id = request.POST.get('message_id')
        if message_id:
            message = Message.objects.filter(id=message_id, receiver=request.user).first()
            if message:
                message.is_read = True
                message.save()
                return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


 

def search_users(request):
    if request.method == "POST":
        query = request.POST.get('q', '').strip()
        print(query,"query")
        name=Group.objects.all()
        print(name)

        # Filter users and groups based on the query
        users = User.objects.filter(username__icontains=query) if query else User.objects.all()
        groupschatname = Group.objects.filter(name__icontains=query) if query else Group.objects.all()
        print(groupschatname,"llpolo")

        # Attach the last message for each user
        for user in users:
            last_message = Message.objects.filter(
                receiver=user
            ).order_by('-created_at').first()
            user.last_message = last_message

        # Check if the request is AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Render partial template for the response
            html = render_to_string('partials/user_list.html', {
                'users': users,
                'groupschatname': groupschatname,
            })
            return JsonResponse({'html': html})
    
    # Default case for GET requests
    users = User.objects.all()
    groupschat = Group.objects.all()
    print(groupschatname,"kk")

    # Attach the last message for each user
    for user in users:
        last_message = Message.objects.filter(
            receiver=user
        ).order_by('-created_at').first()
        user.last_message = last_message

    return render(request, 'sms.html', {'users': users, 'groupschatname': groupschatname})


@login_required
def charget_notifications(request):
    notifications = []
    current_user = request.user

    # Fetch unread notifications for the current user
    unread_notifications = chatNotification.objects.filter(user=current_user, is_read=False)

    for notification in unread_notifications:
        notifications.append({
            'id': notification.id,
            'message_id': notification.message.id,
            'user_id': notification.user.id,
            'username': notification.message.sender.username,  # Add sender's username to notification
            'unread_count': 1,  # Single message, hence 1 unread message
            'type': 'user',
        })

    return JsonResponse({'notifications': notifications})


@csrf_exempt
@login_required
def chat_mark_notification_as_read(request, notification_id):
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'success': False}, status=404)


@login_required
@csrf_exempt
def delete_message(request, message_id):
    print("delete fun")
    if request.method == 'POST':
        message = get_object_or_404(Message, id=message_id, sender=request.user)
        message.delete()
        return JsonResponse({'success': True, 'message': 'Message deleted successfully'})
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


def msg(req):
    return render(req,"send_message.html")



@login_required
@csrf_exempt  # Use this only if CSRF tokens are not being handled properly (like in external APIs)
def create_group(request):
    if request.method == 'POST':
        group_name = request.POST.get('group_name')
        selected_users = request.POST.getlist('level1_users')

        if group_name and selected_users:
            try:
                # Create the group
                group = Group.objects.create(name=group_name, created_by=request.user)

                # Add the creator as a member in the GroupMembership table
                GroupMembership.objects.create(group=group, user=request.user)

                # Add selected users
                for user_id in selected_users:
                    try:
                        user = User.objects.get(id=user_id)
                        GroupMembership.objects.create(group=group, user=user)
                        group.members.add(user)  # Add to group
                        send_group_invitation(user, group)  # Send email
                    except User.DoesNotExist:
                        continue

                return JsonResponse({'status': 'success', 'message': 'Group created successfully!'})

            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        return JsonResponse({'status': 'error', 'message': 'users are required!'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



         
def send_group_invitation(user, group):
    # Create the email content
    message_content = f"You've been added to the group '{group.name}'. Welcome!"
    subject = f"You're added to the group '{group.name}'"
    from_email = settings.DEFAULT_FROM_EMAIL  # You can set this in your settings.py

    # Send the email to the user
    try:
        send_mail(
            subject,
            message_content,
            from_email,
            [user.email],  # List of recipients
            fail_silently=False,  # This will raise an error if email fails
        )
    except Exception as e:
        # print(f"Error sending email to {user.email}: {e}")
        return HttpResponse(f"Error sending email to {user.email}: {e}")




def sms(req):
   
    current_user = req.user  # Get the logged-in user
    user_groups = Group.objects.filter(members=req.user)

    groups = Group.objects.filter(created_by=current_user).order_by('-id').first()  # Adjust ordering as needed
    print("this is grps",groups)
    allgroups = Group.objects.filter(created_by=current_user).order_by('-id')  # Adjust ordering as needed
    # print(allgroups,"all")

   # Retrieve members of the first group
    group_members = groups.members.all() if groups else []  # Fetch members if a group exists
    print("Members of the group:", group_members)
    group_creator = groups.created_by if groups else None  # Fetch the creator if a group exists




    # Get users excluding the current user and annotate with last message time, unread count
    users = User.objects.exclude(id=current_user.id).annotate(
        last_message_time=Max('received_messages__created_at'),  # Most recent received message time
        unread_count=Count('received_messages', filter=Q(received_messages__is_read=False, received_messages__sender=current_user))  # Unread message count
    ).order_by('-last_message_time')  # Sort users by the most recent message time

    # Fetch the actual last message for each user and its content
    for user in users:
        last_message = Message.objects.filter(
            Q(sender=current_user, receiver=user) | Q(sender=user, receiver=current_user)
        ).order_by('-created_at').first()

        # Assign last message and its content
        user.last_message = last_message
        user.last_message_content = last_message.content if last_message else None
        user.last_message_is_read = last_message.is_read if last_message else None
        user.is_online = chatLoginHistory.is_user_online(user)


    # Get the first user from the queryset, or None if the queryset is empty
    first_user = users.first() if users.exists() else None
    # total_users_in_group_count = GroupMembership.objects.filter(group=group).count()
    # group_members = GroupMembership.objects.filter(group=group)

    context = {"users": users, "current_user": current_user, "recipient": first_user,"groups": groups,"allgroupss":user_groups,"group_members":group_members,"group_creator":group_creator,"allgroups":allgroups  # Pass groups to the template
        

    }

    

    return render(req, "sms.html", context)






def generate_meeting_topic():
    # Get the last event in the database
    last_event = Event.objects.order_by('id').last()

    # If no event exists, start with "Meeting-0001"
    if not last_event:
        return "Meeting-0001"

    # Extract the topic from the last event
    last_topic = last_event.topic

    # Ensure the topic follows the expected format
    try:
        # Split the topic and extract the number part
        number = int(last_topic.split('-')[1]) + 1
    except (IndexError, ValueError):
        # Handle cases where the format is unexpected
        return "Meeting-0001"

    # Return the new topic with zero-padded number
    return f"Meeting-{number:04d}"


def meeting(request):
    topic = generate_meeting_topic()
    print("Generated Topic:", topic)  # Debugging line
    users = User.objects.all()  # Fetch all users
    meetings = Event.objects.all()


    print(users,"oopo")


    # Ensure the topic is passed correctly to the template
    return render(request, 'meeting_list.html', {'topic': topic, 'users': users,"meetings":meetings})

def meetingsave(request):
    if request.method == "POST":
        topic = request.POST.get('topic')
        organiser = request.POST.get('organiser')
        partner = request.POST.get('partner')
        partner_logo = request.POST.get('partner_logo')
        event_type = request.POST.get('type')
        participants = request.POST.getlist('participants')  # Get list of selected users
        location = request.POST.get('location')
        event_date = request.POST.get('date')  # Use this as both from_date and to_date
        starttime = request.POST.get('starttime')
        endtime = request.POST.get('endtime')
        link = request.POST.get('link')
        agenda = request.POST.getlist('agenda[]')
        priority = request.POST.get('priority')


        # Create a Project for the meeting
        # default_user = get_object_or_404(User, username='Punithan')
        project = Project(
            projectname=topic,
            priority=priority,
            from_date=event_date,  # Use Event Date for from_date
            to_date=event_date,    # Use Event Date for to_date
            assigned_by=request.user,
            user=None,
        )
        project.save()

         # Create the Event instance
        event = Event(
            topic=topic,
            organiser=organiser,
            partner=partner,
            partner_logo=partner_logo,
            event_type=event_type,
            participants=",".join(participants),  # Store as comma-separated string
            location=location,
            date=event_date,
            starttime=starttime,
            endtime=endtime,
            link=link,
            agenda=agenda,
            project=project,
            prepared_by=request.user,
        )
        event.save()

        return redirect('meeting_list')

    return render(request, 'meeting.html')

def calculate_time_duration(start_time, end_time):
    """
    Calculate the duration between two times, accounting for cross-midnight scenarios.
    """
    start_datetime = datetime.combine(datetime.min, start_time)
    end_datetime = datetime.combine(datetime.min, end_time)
    if end_datetime < start_datetime:
        end_datetime += timedelta(days=1)
    return end_datetime - start_datetime

def meeting_list(request):
    meetings = Event.objects.all()
    print(meetings,"saved meeting")
    users = User.objects.all()  # Fetch all users
    topic = generate_meeting_topic()



    return render(request, 'meeting_list.html', {'meetings': meetings,"users":users,"topic":topic})

#sends the meeting invitation using sending.html
def meetingsend(request, id):
    try:
        # Retrieve the event using the given id
        event = Event.objects.get(id=id)
    except Event.DoesNotExist:
        return render(request, 'meeting.html', {"error": "Event not found"})

    try:
        # Extract participant IDs and filter users accordingly
        participant_ids = [int(pid) for pid in event.participants.split(",")]
    except ValueError:
        return render(request, 'meeting.html', {"error": "Invalid participant IDs format"})

    users = User.objects.filter(id__in=participant_ids)

    # Create a list of email addresses for the participants
    to_email = [user.email for user in users]

    # Prepare context for the email body
    context = {
        'topic': event.topic,
        'organiser': event.organiser,
        'partner': event.partner,
        'partner_logo': event.partner_logo,
        'type': event.event_type,
        'location': event.location,
        'date': event.date,
        'starttime': event.starttime,
        'endtime': event.endtime,
        'duration': event.duration,
        'link': event.link,
        'agenda': event.agenda,
    }

    # Render the HTML content for the email
    html_content = render_to_string('sending.html', context)

    # Email details
    subject = f"Meeting Invitation: {event.topic}"
    from_email = "taskaccsys@gmail.com"
    
    # Sending the email to the participants
    email = EmailMultiAlternatives(subject, "", from_email, to_email)
    email.attach_alternative(html_content, "text/html")
    
    try:
        # Send the email to all participants
        email.send()
    except Exception as e:
        # Handle error if sending fails
        return render(request, 'meeting.html', {"error": f"Error sending email: {str(e)}"})

    # Redirect to the meeting list page after sending the email
    return redirect('meeting_list')

def after_meeting(request, id):
    meeting = Event.objects.get(id=id)
    return render(request, 'add_remark.html', {'meeting': meeting})

def delete_meeting(request, id):

    meeting = get_object_or_404(Event, id=id)
    meeting.delete()
    messages.success(request, 'Meeting deleted successfully!')
    return redirect('meeting_list')

def points_discussed(request, id):
    if request.method == 'POST':
        meeting = get_object_or_404(Event, id=id)

        meeting.actual_starttime  = request.POST.get('actual_starttime')
        meeting.actual_endtime = request.POST.get('actual_endtime')
        remarks = request.POST.getlist('remark[]')
        try:
            formatted_actual_start_time = datetime.strptime(meeting.actual_starttime, "%H:%M").time()
            formatted_actual_end_time = datetime.strptime(meeting.actual_endtime, "%H:%M").time()
        except ValueError:
            return render(request, 'meeting.html', {"error": "Invalid time format. Use HH:MM."}) 
        
        meeting.actual_duration = calculate_time_duration(formatted_actual_start_time, formatted_actual_end_time)

        filtered_remarks = [remark.strip() for remark in remarks if remark.strip()] 
        meeting.remark = filtered_remarks  
        meeting.save()

        return redirect('meeting_list')

def points_agreed(request, id):
    meeting = get_object_or_404(Event, id=id)

    if request.method == 'POST':
        # Retrieve form data
        remark = request.POST.get('remark')
        selected_user_ids = request.POST.getlist('selected_users')
        priority = request.POST.get('priority', 'Low')
        assigned_date = request.POST.get('assigned_date')
        final_date = request.POST.get('final_date')
        description = request.POST.get('description', '')

        # Convert dates
        try:
            assigned_date = datetime.strptime(assigned_date, "%Y-%m-%d").date()
            final_date = datetime.strptime(final_date, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

        # Get selected users
        selected_users = User.objects.filter(id__in=selected_user_ids)

        if not selected_users:
            messages.error(request, "Please select at least one user to assign the task.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Fetch the linked project
        project = meeting.project
        if not project:
            messages.error(request, "No project linked to this meeting.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Create tasks and send notifications
        for user in selected_users:
            # Create the task
            task = Task.objects.create(
                taskname=remark,
                priority=priority,
                from_date=assigned_date,
                to_date=final_date,
                description=description,
                user=user,
                project=project,
                assigned_by=request.user
            )

            # Notify user in-app
            Notification.objects.create(
                user=user,
                message=f"Task '{remark}' created by {request.user.username}",
                assigned_by=request.user
            )

            # Send email notification
            if user.email:
                subject = f"New Task Assigned: {remark}"
                message = (
                    f"Dear {user.username},\n\n"
                    f"You have been assigned a new task in the project '{project.projectname}'.\n\n"
                    f"Task Name: {remark}\n"
                    f"Priority: {priority}\n"
                    f"Assigned Date: {assigned_date}\n"
                    f"Final Date: {final_date}\n"
                    f"Description: {description}\n\n"
                    f"Best regards,\n"
                    f"{request.user.username}"
                )
                try:
                    send_mail(subject, message, settings.EMAIL_HOST_USER, [user.email], fail_silently=False)
                except Exception as e:
                    # Log or handle email-sending error
                    messages.error(request, f"Failed to send email to {user.username}: {str(e)}")

        # Redirect back to the referring page
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # GET request: Render the form
    assigned_remarks = [task.taskname for task in meeting.project.tasks.all()] if meeting.project else []
     # Parse participants into a list of user IDs
    participants_ids = list(map(int, meeting.participants.split(','))) if meeting.participants else []

    # Fetch users from the participants
    participants = User.objects.filter(id__in=participants_ids).exclude(groups__name="Superadmin")

    return render(request, 'assign_tasks.html', {
        'meeting': meeting,
        'remarks': meeting.remark,
        'assigned_remarks': assigned_remarks,
        'users': participants  
    })

def minutes_of_meeting(request, id):
    # Fetch the Event object
    meeting = get_object_or_404(Event, id=id)
    
    # Fetch tasks related to the project's tasks
    tasks = Task.objects.filter(project=meeting.project)  # Use the project to filter tasks
    
    # Convert participant IDs into user names
    participant_ids = meeting.participants.split(",")  # Split IDs stored as "1,2,12"
    participants = User.objects.filter(id__in=participant_ids)  # Fetch user objects
    
    # Pass the data to the template
    return render(request, 'minutes_of_meeting.html', {
        'meeting': meeting,
        'tasks': tasks,
        'participants': participants,
    })

def send_mom(request, id):
    # Fetch the Event object
    meeting = get_object_or_404(Event, id=id)
    
    # Fetch associated tasks from the Task table
    tasks = Task.objects.filter(project=meeting.project)  # Use project to filter tasks

    # Get participants' email addresses
    participant_ids = meeting.participants.split(",")  # Split participant IDs into a list
    participants = User.objects.filter(id__in=participant_ids)  # Fetch user objects
    print("participants",participants)
    to_email = [user.email for user in participants if user.email]  # Get their email addresses

    if not to_email:
        return HttpResponse("No participants with valid email addresses found.")

    # Prepare email context
    context = {
        'meeting': {
            'topic': meeting.topic,
            'organiser': meeting.organiser,
            'partner': meeting.partner,
            'event_type': meeting.event_type,
            'participants': participants,
            'location': meeting.location,
            'date': meeting.date,
            'starttime': meeting.starttime,
            'endtime': meeting.endtime,
            'actual_starttime': meeting.actual_starttime,
            'actual_endtime': meeting.actual_endtime,
            'duration': meeting.duration,
            'actual_duration': meeting.actual_duration,
            'agenda': meeting.agenda,
            'remark': meeting.remark,
            'link': meeting.link,
        },
        'tasks': tasks,
    }

    # Render the HTML content
    html_content = render_to_string('sending_mom.html', context)

    # Prepare email
    subject = f"Minutes of the Meeting - {meeting.topic}"
    from_email = "taskaccsys@gmail.com"  # Replace with your email
    email = EmailMultiAlternatives(subject, "", from_email, to_email)
    email.attach_alternative(html_content, "text/html")

    # Send the email
    try:
        email.send()
        return redirect('meeting_list')
    except Exception as e:
        return HttpResponse(f"Failed to send email: {str(e)}")

def workflowmgt(request):
    return render(request, 'components.html')  



 

class LeaveRequestView(View):
    allowed_file_types = ['pdf', 'xls', 'xlsx', 'doc', 'docx', 'ppt', 'pptx', 'jpg', 'jpeg', 'png', 'txt']

    def get(self, request, *args, **kwargs):
        user_leave_requests = LeaveRequest.objects.filter(user=request.user).order_by('-created_at')

        pending_approval_requests = LeaveRequest.objects.filter(
            status="Pending",
            current_level__in=[1, 2, 3]
        ).exclude(user=request.user).filter(
            Q(level1_approvers=request.user, current_level=1, status="Pending") |
            Q(level2_approvers=request.user, current_level=2, status="Pending") |
            Q(level3_approvers=request.user, current_level=3, status="Pending")
        ).distinct().order_by('-created_at')

        for leave_request in user_leave_requests:
            if leave_request.from_date and leave_request.to_date:
                leave_request.no_of_days = (leave_request.to_date - leave_request.from_date).days + 1
            else:
                leave_request.no_of_days = 0

        for leave_request in pending_approval_requests:
            if leave_request.from_date and leave_request.to_date:
                leave_request.no_of_days = (leave_request.to_date - leave_request.from_date).days + 1
            else:
                leave_request.no_of_days = 0

        context = {
            'user_leave_requests': user_leave_requests,
            'pending_approval_requests': pending_approval_requests,
            'level1_users': User.objects.filter(groups__name='Level1'),
            'level2_users': User.objects.filter(groups__name='Level2'),
            'level3_users': User.objects.filter(groups__name='Level3'),  
                    }

        return render(request, 'leavedetails.html', context)
        
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        max_file_size = 5 * 1024 * 1024  
        
        if file:
            extension = os.path.splitext(file.name)[1][1:].lower()
            if extension not in self.allowed_file_types:
                return JsonResponse({
                    'status': 'error',
                    'message': f"File type not allowed. Allowed types: {', '.join(self.allowed_file_types)}"
                })
            
        if file!=None:

            if file.size > max_file_size:
                return JsonResponse({
                    'status': 'error',
                    'message': "File size exceeds the 5 MB limit."
                })

        leave_request = LeaveRequest.objects.create(
            user=request.user,
            leave_type=request.POST.get('LeaveType'),
            from_date=request.POST.get('fromdate'),
            to_date=request.POST.get('todate'),
            session_from=request.POST.get('Session1'),
            session_to=request.POST.get('Session2'),
            reason=request.POST.get('Reason'),
            file=file,
            status='Pending',
            created_at=timezone.now(),
            current_level=1
        )

        for level, field_name in enumerate(['level1_users[]', 'level2_users[]', 'level3_users[]'], start=1):
            user_ids = request.POST.getlist(field_name)
            users = User.objects.filter(id__in=user_ids)
            getattr(leave_request, f'level{level}_approvers').set(users)
            if level == 1:  
                self.notify_approvers(users, leave_request)

        return JsonResponse({'status': 'success', 'message': 'Leave applied successfully!'})

    def notify_approvers(self, approvers, leave_request):
        subject = f"Leave Request Approval Needed from {leave_request.user.username}"
        message = (
            f"A leave request from {leave_request.user.username} requires your approval.\n\n"
            f"Leave Type: {leave_request.leave_type}\n"
            f"From: {leave_request.from_date}\n"
            f"To: {leave_request.to_date}\n"
            f"Reason: {leave_request.reason}\n"
        )
        email = EmailMessage(
            subject,
            message,
            settings.EMAIL_HOST_USER,
            [approver.email for approver in approvers]

        )

        if leave_request.file:
            leave_request.file.open('rb')  
            mime_type, _ = mimetypes.guess_type(leave_request.file.name)
            email.attach(leave_request.file.name, leave_request.file.read(), mime_type or 'application/octet-stream')
    
        try:
            email.send(fail_silently=False)

        except Exception as e:
            print(f"Failed to send email: {e}")
            return JsonResponse("Networks error please try again later")

@csrf_exempt
def approve_leave(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id)
    approvers, next_level_users = None, None
    
    # Determine current level approvers and next level approvers
    if leave_request.current_level == 1:
        approvers = leave_request.level1_approvers.all()
        next_level_users = leave_request.level2_approvers.all()
    elif leave_request.current_level == 2:
        approvers = leave_request.level2_approvers.all()
        next_level_users = leave_request.level3_approvers.all()
    elif leave_request.current_level == 3:
        approvers = leave_request.level3_approvers.all()

    # Check if the user is authorized to approve at the current level
    if request.user not in approvers:
        return JsonResponse({'status': 'error', 'message': 'You are not authorized to approve this request.'})

    # Update leave request's current level or mark as Approved if at the final level
    if leave_request.current_level < 3:
        leave_request.current_level += 1
        leave_request.status = 'Pending'
        leave_request.save()
        
        # Notify the next level approvers if there are any
        if next_level_users:
            LeaveRequestView().notify_approvers(next_level_users, leave_request)
        messages.success(request, "Leave request moved to the next approval level.")
    else:
        # Mark as Approved when all levels have approved
        leave_request.status = 'Approved'
        leave_request.save()
        messages.success(request, "Leave request approved successfully.")

    return JsonResponse({'status': 'success', 'message': 'Leave request approved successfully.'})



def reject_leave(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id)

    if request.method == 'POST':
        # Get the rejection reason from the request data (from the JavaScript AJAX POST request)
        data = json.loads(request.body)  # Extracting data sent by the frontend (JSON)
        reason = data.get('Reason')

        if reason:
            # Save the reason and update the leave status to 'Rejected'
            leave_request.rejection_reason = reason  # Assuming 'reason' field exists in the model
            leave_request.status = 'Rejected'
            leave_request.save()

            

            # Add a success message to notify that the leave request was rejected
            messages.success(request, "Leave request rejected successfully.")
            print("leave rejected")
             # Send email notification to the user
            subject = "Your Leave Request has been Rejected"
            message = f"Dear {leave_request.user.username},\n\nYour leave request has been rejected.\nReason: {reason}\n\nThank you."
            recipient = leave_request.user.email

            try:
                send_mail(subject, message, settings.EMAIL_HOST_USER, [recipient])
            except Exception as e:
                print(f"Failed to send email: {e}")
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'Reason is required'})
      # Handle GET request (show the leave request details)
  


class WithdrawLeaveView(LoginRequiredMixin, View):
    
    def post(self, request, leave_id):
        print("Withdrawal function called")
        
        # Fetch the leave request based on the leave_id
        leave_request = get_object_or_404(LeaveRequest, id=leave_id, user=request.user)
        print("Processing POST request")

        # Delete the leave request
        leave_request.delete()

        return JsonResponse({'status': 'success', 'message': 'Leave request withdrawn successfully.'})
    
    def get(self, request, leave_id):
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})
    


def gst1(req):
    return render(req,"gst.html")



def download_sample_book(request):
    # Create a workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sample Book"

    # Add headers to the sheet
    headers = [
        "Invoice Number", "Invoice Date", "GSTIN", "Customer Name","Ledger Name","Invoice Type",
        "GST Rate","HSN","Taxable","IGST", "CGST", "SGST",]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the headers bold

    # Set response headers
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="book.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response



def download_sample_portal(request):
    # Create a workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sample Portal"

    # Add headers to the sheet
    headers = [
        "Invoice Number", "Invoice Date", "GSTIN", "Customer Name","GST Rate","HSN","Taxable","IGST", "CGST","SGST",   
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the headers bold

    # Set response headers
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Portal.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response

def mailtemp(req):
    return render(req,"send_email.html")


# Function to remove HTML tags
def strip_html_tags(text):
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text)


# @login_required
# def send_message(request):
#     print("send msg")
#     if request.method == "POST":
#         recipient_email = request.POST.get("to_email")
#         cc_emails = request.POST.get("cc")  # Retrieve CC emails
#         subject = request.POST.get("subject")
#         body = request.POST.get("message")  # Retrieve TinyMCE field value
#         attached_file = request.FILES.get("file")  # Get the uploaded file

#         print(body, "body")
#         print(recipient_email, "email")
#         plain_body = strip_html_tags(body)

#         # Validate recipient email format
#         if not re.match(r"[^@]+@[^@]+\.[^@]+", recipient_email):
#             return JsonResponse({
#                 "status": "error",
#                 "message": "Invalid recipient email address."
#             })
        
#         # Validate CC emails
#         if cc_emails:
#             cc_email_list = [email.strip() for email in cc_emails.split(",")]
#             invalid_emails = [email for email in cc_email_list if not re.match(r"[^@]+@[^@]+\.[^@]+", email)]
#             if invalid_emails:
#                 return JsonResponse({
#                     "status": "error",
#                     "message": f"Invalid email address(es) in CC: {', '.join(invalid_emails)}"
#                 })
#         else:
#             cc_email_list = []

#         # Save the message (receiver can be None since it's an arbitrary email)
#         message = Messagesytem.objects.create(
#             sender=request.user,
#             receiver=None,  # No linked User instance
#             recipient_email=recipient_email,  # Save the provided email address

#             subject=subject,
#             content=plain_body,  # Save TinyMCE content
#             file=attached_file,
#         )

#         print(message.content, "this is body")

#         try:
#             # Set the Reply-To address to the sender's email
#             message_id = make_msgid()  # Gmail will replace the domain automatically
#             reply_to_email = request.user.email
#             print(reply_to_email, "reply mail")

#             # Send the email
#             email = EmailMessage(
#                 subject=subject,
#                 body=plain_body,
#                 from_email=request.user.email,
#                 to=[recipient_email],  # Send directly to the arbitrary email address
#                 cc=cc_email_list,
#                 reply_to=[reply_to_email],  # This ensures replies go to the sender
#                 headers={'Message-ID': message_id},  # Add the Message-ID header
#             )

#             if attached_file:
#                 email.attach(attached_file.name, attached_file.read(), attached_file.content_type)

#             email.send()

#             # Update the message with Message-ID
#             message.message_id = message_id
#             message.save()

#             return JsonResponse({
#                 "status": "success",
#                 "message": "Email sent successfully!"
#             })

#         except Exception as e:
#             return JsonResponse({
#                 "status": "error",
#                 "message": f"Failed to send email: {str(e)}"
#             })

#     users = User.objects.exclude(id=request.user.id)
#     return render(request, 'send_message.html', {"users": users})


# @login_required
# def inboxes(request):
#     messages = Messagesytem.objects.filter(receiver=request.user)
#     return render(request, 'inboxes.html', {'messages': messages})







# @login_required
# def get_notifications(request):
#     # Get unread and unshown notifications
#     notifications = Notification.objects.filter(recipient=request.user, is_read=False, is_shown=False)
#     notification_count = notifications.count()

#     # Mark notifications as shown
#     notifications.update(is_shown=True)

#     return JsonResponse({"count": notification_count})

# @login_required
# def check_new_notifications(request):
#     # Fetch notifications that are unread and not shown
#     unread_notifications = Notification.objects.filter(
#         recipient=request.user, is_read=False
#     )

#     if unread_notifications.exists():
#         # Get the notification count
#         notification_count = unread_notifications.count()

#         # Prepare response data
#         notification_data = {
#             "new_notifications": True,
#             "message": f"You have {notification_count} new message(s)!",
#         }

#         # Mark the notifications as read when they are fetched
#         unread_notifications.update(is_read=True)
#     else:
#         notification_data = {"new_notifications": False}

#     return JsonResponse(notification_data)


# def outlook(req):
#     messages = Messagesytem.objects.filter(sender=req.user)

#     return render(req, 'mail.html', {'messages': messages})



# def delete_message(request, message_id):
#     if request.method == "POST":
#         # Retrieve the message by ID
#         message = get_object_or_404(Messagesytem, id=message_id, sender=request.user)

#         # Delete the message
#         message.delete()

#         # Check if the request is an AJAX request
#         if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#             return JsonResponse({"status": "success", "message": "Message deleted successfully."})

#         # Redirect for non-AJAX requests
#         return JsonResponse({"status": "error", "message": "Invalid request."})

#     # If not POST, return an error
#     return JsonResponse({"status": "error", "message": "Invalid request."})




@login_required
def get_stored_emails(request):
    emails = list(User.objects.values_list('email', flat=True))  # Replace 'User' with your model if different
    return JsonResponse({'emails': emails})






# ---------------------------------------------------------------------------------

# views.py


def maill(req):
    return render(req,"send_email.html")


# # Ensure that the user is logged in before sending an email
# @login_required
# def send_email_view(request):
#     if request.method == "POST":
#         print("post")
#         subject = request.POST.get("subject")
#         body = request.POST.get("body")
#         recipient = request.POST.get("recipient")

#         # Dynamically use the currently logged-in user's email as the sender
#         sender = request.user.email
        
#         # Send the email using Django's built-in send_mail function
#         send_mail(subject, body, sender, [recipient])
#         mailbox = Mailbox.objects.first()  # Retrieve the first mailbox (or specify one)


#         # Create a Message record for the sent email (you can exclude from_address as it's managed by the library)
#         # You can store the email subject, body, and recipient in the Message model if needed.
#         message=Message.objects.create(
#             subject=subject,
#             body=body,
#             mailbox=mailbox,  # Associate the message with the mailbox
#             from_header=sender,  # Explicitly set the sender email
#             to_header=recipient,  # Explicitly set the recipient email
#             processed=now(), 

#             # Optionally, you can add other fields like 'from_address', 'message_id', etc.
#             # Don't manually set 'from_address' here since it's handled internally
#         )

#         # message.save()
        
#         # return redirect('show_replied_email', email_id=message.id)
#         return HttpResponse("success")
#     else:
#             print("else")
#             return render(request, "send_email.html")

        
    
#     # return render(request, "send_email.html")

# def sent_emails_view(request):
#     user_email = request.user.email
#     # Debug: Print the email address
#     print(f"Logged-in user email: {user_email}")
    
#     # Attempt to match emails more robustly
#     sent_emails = Message.objects.filter(from_header__icontains=user_email).order_by('-processed')
    
#     # Debug: Log the query results
#     print(f"Sent emails: {sent_emails}")
    
#     return render(request, "sent_emails.html", {"sent_emails": sent_emails})




# # views.py (continuation)
# def inbox_view(request):
#     # Fetch all messages from the inbox (these messages come from Mailbox)
#     inbox_messages = Message.objects.filter(mailbox__is_active=True).order_by('-received_at')
#     return render(request, "inbox.html", {"messages": inbox_messages})


# def received_emails_view(request):
#     # Get all received and replied emails
#     received_emails = CustomMessage.objects.filter(read=True)
#     replied_emails = CustomMessage.objects.filter(original_message__isnull=False)

#     # Pass emails to the template
#     return render(request, "received_emails.html", {
#         "received_emails": received_emails,
#         "replied_emails": replied_emails
#     })# def process_incoming_emails():
# #     print("process income ing c=mails")
# #     for message in Message.objects.filter(processed=False):  # Unprocessed messages
# #         print(f"Processing message ID: {message.id}")
# #         original_email_id = message.in_reply_to
# #         if original_email_id:
# #             original_message = Message.objects.filter(message_id=original_email_id).first()
# #             if original_message:
# #                 original_message.replies.add(message)
# #         message.processed = True
# #         message.save()





